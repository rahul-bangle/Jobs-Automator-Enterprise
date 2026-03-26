from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlmodel import select
from app.core.db import get_session
from app.models.base import Job, ImportBatch
import csv
import io
import openpyxl
from typing import List

router = APIRouter()


def _parse_row(row: dict) -> dict | None:
    """Normalize flexible CSV/XLSX column names to standard fields."""
    company = (
        row.get("Company") or row.get("company") or
        row.get("company_name") or row.get("Company Name")
    )
    title = (
        row.get("Title") or row.get("title") or
        row.get("job_title") or row.get("Job Title")
    )
    url = (
        row.get("URL") or row.get("url") or
        row.get("source_url") or row.get("Job URL") or row.get("Link")
    )
    location = (
        row.get("Location") or row.get("location") or
        row.get("City") or row.get("city")
    )

    if not all([company, title, url]):
        return None  # Reject incomplete rows

    return {
        "company_name": str(company).strip(),
        "job_title": str(title).strip(),
        "source_url": str(url).strip(),
        "location": str(location).strip() if location else "Remote / TBD"
    }


@router.post("/csv")
async def import_csv(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session)
):
    """Import jobs from a CSV file. Accepts flexible column names."""
    content = await file.read()
    decoded = content.decode("utf-8-sig")  # Handle BOM-encoded CSVs
    reader = csv.DictReader(io.StringIO(decoded))

    batch = ImportBatch(source_type="CSV", file_name=file.filename)
    session.add(batch)
    await session.flush()

    accepted, skipped = [], 0

    for raw_row in reader:
        parsed = _parse_row(dict(raw_row))
        if not parsed:
            skipped += 1
            continue

        job_id = Job.generate_id(parsed["company_name"], parsed["job_title"], parsed["location"])
        existing = await session.get(Job, job_id)
        if existing:
            skipped += 1
            continue

        job = Job(id=job_id, queue_status="review", **parsed)
        session.add(job)
        accepted.append(job_id)

    batch.total_rows = len(accepted) + skipped
    batch.accepted_rows = len(accepted)
    batch.rejected_rows = skipped
    await session.commit()

    return {
        "message": f"Imported {len(accepted)} jobs",
        "batch_id": batch.id,
        "accepted": len(accepted),
        "skipped": skipped
    }


@router.post("/xlsx")
async def import_xlsx(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session)
):
    """Import jobs from an XLSX file."""
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [str(cell.value).strip() for cell in ws[1]]
    batch = ImportBatch(source_type="XLSX", file_name=file.filename)
    session.add(batch)
    await session.flush()

    accepted, skipped = [], 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_row = dict(zip(headers, [str(v) if v is not None else "" for v in row]))
        parsed = _parse_row(raw_row)
        if not parsed:
            skipped += 1
            continue

        job_id = Job.generate_id(parsed["company_name"], parsed["job_title"], parsed["location"])
        existing = await session.get(Job, job_id)
        if existing:
            skipped += 1
            continue

        job = Job(id=job_id, queue_status="review", **parsed)
        session.add(job)
        accepted.append(job_id)

    batch.total_rows = len(accepted) + skipped
    batch.accepted_rows = len(accepted)
    batch.rejected_rows = skipped
    await session.commit()

    return {
        "message": f"Imported {len(accepted)} jobs",
        "batch_id": batch.id,
        "accepted": len(accepted),
        "skipped": skipped
    }


from pydantic import BaseModel
from typing import List, Optional
from app.services.pipeline_v2 import discovery_service
from app.services.scoring import scoring_service

class PreviewRequest(BaseModel):
    urls: str
    fileName: Optional[str] = None

@router.post("/preview")
async def preview_import_urls(
    payload: PreviewRequest,
    session: AsyncSession = Depends(get_session)
):
    """
    Live scraping and AI scoring for the frontend preview.
    Uses Auto-Fetch Lite (BS4/httpx fallback) and Groq.
    """
    urls = [url.strip() for url in payload.urls.splitlines() if url.strip()]
    rows = []
    
    for url in urls:
        try:
            # 1. Scraping (Auto-Fetch Lite)
            html = await discovery_service.fetch_url(url)
            content = await discovery_service.process_html(html)
            # 2. AI Scoring (Groq)
            score_result = await scoring_service.score_job(
                session=session,
                job_title="Product Manager", # Ideally extract from content
                company_name="Scraped Co",    # Ideally extract from content
                location="Remote",
                description=content,
                candidate_profile="Product Manager with experience in AI and Ops."
            )
            rows.append({
                "id": f"preview-{Job.generate_id('Scraped', url, 'Remote')}",
                "companyName": "Scraped Listing", # In future, extract from content
                "jobTitle": "Product Manager",    # In future, extract from content
                "source": "Auto-Fetch",
                "sourceUrl": url,
                "location": "Remote",
                "trustScore": 85, # Default for verified scraping
                "relevanceScore": int(score_result["overall_score"] * 100),
                "queueStatus": "accepted" if score_result["overall_score"] > 0.8 else "review",
                "duplicate": False,
                "validationError": ""
            })
        except Exception as e:
            rows.append({
                "id": f"error-{hash(url)}",
                "companyName": "Failed to Fetch",
                "jobTitle": "N/A",
                "source": "Error",
                "sourceUrl": url,
                "location": "N/A",
                "trustScore": 0,
                "relevanceScore": 0,
                "queueStatus": "rejected",
                "duplicate": False,
                "validationError": str(e)
            })

    return {
        "fileName": payload.fileName,
        "rows": rows,
        "summary": {
            "accepted": len([r for r in rows if r["queueStatus"] == "accepted"]),
            "review": len([r for r in rows if r["queueStatus"] == "review"]),
            "rejected": len([r for r in rows if r["queueStatus"] == "rejected"]),
            "duplicates": 0
        }
    }
